from __future__ import annotations

import csv
from datetime import UTC, datetime
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from app.nutrition.application.food_evidence_candidate_record_values import stable_hash
from app.nutrition.application.food_evidence_candidate_source_file_records import (
    csv_records,
    find_source_file,
    json_records,
)
from app.nutrition.application.food_evidence_candidate_source_file_records import (
    schema_keys as collect_schema_keys,
)
from app.nutrition.application.food_evidence_candidate_source_field_mapping import record_id
from app.nutrition.application.food_evidence_candidate_xlsx_records import (
    detect_tfda_header,
    xlsx_row_record,
)
from app.nutrition.application.food_raw_source_inventory import (
    NON_CLAIM_FLAGS,
    RAW_SOURCE_DEFINITIONS,
    RawSourceDefinition,
    pipeline_stage_boundary,
)

NO_TRUTH_FLAGS = {
    **NON_CLAIM_FLAGS,
    "runtime_truth_changed": False,
}


def build_food_raw_source_vault(scan_roots: Iterable[Path | str]) -> dict[str, Any]:
    roots = [Path(root) for root in scan_roots]
    reports: list[dict[str, Any]] = []
    raw_records: list[dict[str, Any]] = []

    for definition in RAW_SOURCE_DEFINITIONS:
        report, records = _ingest_source(definition, roots)
        reports.append(report)
        raw_records.extend(records)

    present_reports = [report for report in reports if report["local_path_present"]]
    return {
        "artifact_type": "accurate_intake_food_raw_source_vault",
        "artifact_schema_version": "1.0",
        "generated_at_utc": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "claim_scope": "raw_source_vault_only",
        "truth_owner": "FoodDB",
        "semantic_owner": "none",
        "runtime_truth": False,
        **NO_TRUTH_FLAGS,
        "pipeline_stage_boundary": {
            "implemented_stage": "raw_source_vault",
            "previous_stage": pipeline_stage_boundary()["implemented_stage"],
            "next_stages_not_implemented": [
                "candidate",
                "validator_passed",
                "auto_eligible_packet_candidate",
                "packet_ready",
            ],
        },
        "vault_summary": {
            "source_count": len(RAW_SOURCE_DEFINITIONS),
            "present_source_count": len(present_reports),
            "raw_record_count": len(raw_records),
            "parse_error_count": sum(1 for report in reports if report["parse_error"]),
        },
        "source_reports": reports,
        "raw_records": raw_records,
    }


def _ingest_source(
    definition: RawSourceDefinition,
    scan_roots: list[Path],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    report = _source_report(definition)
    path = find_source_file(definition.filename, scan_roots)
    if path is None:
        return report, []

    report["local_path_present"] = True
    try:
        records = _records_for_path(path)
    except (csv.Error, json.JSONDecodeError, OSError, UnicodeError) as exc:
        report["parse_error"] = type(exc).__name__
        return report, []

    report.update(
        {
            "parsed_count": len(records),
            "schema_keys": collect_schema_keys(records),
        }
    )
    raw_records = [
        _raw_record(definition, raw_record=record, row_index=index)
        for index, record in enumerate(records, start=_start_row(path))
        if isinstance(record, dict)
    ]
    report["raw_record_count"] = len(raw_records)
    return report, raw_records


def _records_for_path(path: Path) -> list[Any]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return json_records(json.loads(path.read_text(encoding="utf-8-sig")))
    if suffix == ".csv":
        return csv_records(path)
    if suffix == ".xlsx":
        return _xlsx_records(path)
    return []


def _xlsx_records(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return []
    header_index, columns = _detect_xlsx_header(rows)
    return [
        xlsx_row_record(row, columns)
        for row in rows[header_index + 1 :]
        if any(value is not None for value in row)
    ]


def _detect_xlsx_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    first_row = [str(value or "").strip() for value in rows[0]]
    if "label" in first_row and "kcal" in first_row:
        return 0, {
            "category": first_row.index("category") if "category" in first_row else 0,
            "label": first_row.index("label"),
            "aliases": first_row.index("aliases") if "aliases" in first_row else 2,
            "kcal": first_row.index("kcal"),
            "corrected_kcal": first_row.index("corrected_kcal")
            if "corrected_kcal" in first_row
            else first_row.index("kcal"),
        }
    return detect_tfda_header(rows)


def _raw_record(
    definition: RawSourceDefinition,
    *,
    raw_record: dict[str, Any],
    row_index: int,
) -> dict[str, Any]:
    return {
        "record_kind": "raw_source_row",
        "source_id": definition.source_id,
        "source_class": definition.source_class,
        "source_role": definition.source_role,
        "candidate_role": definition.candidate_role,
        "runtime_truth_allowed_default": definition.runtime_truth_allowed_default,
        "macro_truth_allowed_default": definition.macro_truth_allowed_default,
        "source_file": definition.filename,
        "row_index": row_index,
        "record_id": record_id(raw_record),
        "raw_record_hash": stable_hash(raw_record),
        "raw_record": raw_record,
    }


def _source_report(definition: RawSourceDefinition) -> dict[str, Any]:
    return {
        "source_id": definition.source_id,
        "filename": definition.filename,
        "source_class": definition.source_class,
        "candidate_role": definition.candidate_role,
        "local_path_present": False,
        "parsed_count": 0,
        "raw_record_count": 0,
        "schema_keys": [],
        "parse_error": None,
    }


def _start_row(path: Path) -> int:
    if path.suffix.lower() == ".csv":
        return 2
    if path.suffix.lower() == ".xlsx":
        return 2
    return 1


__all__ = ["build_food_raw_source_vault"]
