from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
import csv
import hashlib
import json
from pathlib import Path
from typing import Any, Iterable
from xml.etree.ElementTree import ParseError, iterparse
from zipfile import BadZipFile, ZipFile

from openpyxl.reader.excel import ExcelReader
from openpyxl.utils.exceptions import InvalidFileException


NON_CLAIM_FLAGS = {
    "food_kb_truth_updated": False,
    "nutrition_seed_created": False,
    "exact_card_created": False,
    "packet_truth_created": False,
    "canonical_eval_promoted": False,
}
XLSX_PARSE_EXCEPTIONS = (
    BadZipFile,
    InvalidFileException,
    ParseError,
    OSError,
    ValueError,
    KeyError,
)


@dataclass(frozen=True)
class RawSourceDefinition:
    source_id: str
    filename: str
    source_class: str
    intended_roles: tuple[str, ...]
    source_role: str = "raw_source"
    runtime_truth: bool = False
    notes: str = ""

    def as_dict(self) -> dict[str, Any]:
        return {
            "source_id": self.source_id,
            "filename": self.filename,
            "source_class": self.source_class,
            "source_role": self.source_role,
            "intended_roles": list(self.intended_roles),
            "runtime_truth": self.runtime_truth,
            "notes": self.notes,
        }


RAW_SOURCE_DEFINITIONS: tuple[RawSourceDefinition, ...] = (
    RawSourceDefinition(
        source_id="tfda_fda_food_nutrition_2024",
        filename="FDA_food_nutrition_2024.xlsx",
        source_class="taiwan_tfda_open_data",
        intended_roles=("generic_anchor", "listed_component_anchor"),
        notes="TFDA/FDA nutrition Excel inventory only; not packet truth.",
    ),
    RawSourceDefinition(
        source_id="tfda_tnfcds_consumer_detail",
        filename="tnfcds_consumer_detail.xlsx",
        source_class="taiwan_tfda_open_data",
        intended_roles=("generic_anchor", "listed_component_anchor"),
        notes="TNFCDS consumer detail raw/staging source inventory only.",
    ),
    RawSourceDefinition(
        source_id="tfda_tnfcds_consumer_items",
        filename="tnfcds_consumer_items.xlsx",
        source_class="taiwan_tfda_open_data",
        intended_roles=("generic_anchor", "listed_component_anchor"),
        notes="TNFCDS consumer items raw/staging source inventory only.",
    ),
    RawSourceDefinition(
        source_id="newtaipei_brand_candidates",
        filename="newtaipei_brand_candidates.json",
        source_class="official_brand_chain_page",
        source_role="staging_candidate_only",
        intended_roles=("exact_card_candidate",),
        notes="Official page-derived brand candidates only; no runtime truth.",
    ),
    RawSourceDefinition(
        source_id="local_tw_packaged_extract_188_2",
        filename="188_2.csv",
        source_class="local_taiwan_packaged_extract",
        source_role="staging_candidate_only",
        intended_roles=("exact_card_candidate",),
        notes="Local extracted CSV packaged-product trace only; candidate review only and never runtime truth in this slice.",
    ),
    RawSourceDefinition(
        source_id="openfoodfacts_taiwan_small",
        filename="openfoodfacts_taiwan_small.json",
        source_class="open_food_facts",
        intended_roles=("packaged_candidate",),
        notes="OpenFoodFacts Taiwan sample inventory only.",
    ),
    RawSourceDefinition(
        source_id="usda_food_list_sample",
        filename="usda_food_list_sample.json",
        source_class="usda_fallback",
        intended_roles=("fallback_anchor",),
        notes="USDA fallback sample inventory only.",
    ),
    RawSourceDefinition(
        source_id="base_nutrition_db",
        filename="base_nutrition_db.json",
        source_class="existing_repo_seed",
        source_role="candidate_only",
        intended_roles=("alias_coverage_prior",),
        notes="Existing repo seed used as alias coverage prior only.",
    ),
    RawSourceDefinition(
        source_id="tfda_base_candidates",
        filename="tfda_base_candidates.json",
        source_class="taiwan_tfda_open_data",
        source_role="staging_candidate_only",
        intended_roles=("generic_anchor", "listed_component_anchor"),
        notes="Existing TFDA staging candidates only.",
    ),
    RawSourceDefinition(
        source_id="tfda_base_review_candidates",
        filename="tfda_base_review_candidates.json",
        source_class="taiwan_tfda_open_data",
        source_role="staging_candidate_only",
        intended_roles=("generic_anchor", "listed_component_anchor"),
        notes="Existing TFDA review staging candidates only.",
    ),
)


def build_food_raw_source_registry() -> dict[str, Any]:
    return {
        "artifact_type": "accurate_intake_food_raw_source_registry",
        "artifact_schema_version": "1.0",
        "claim_scope": "raw_source_registry_only",
        "truth_owner": "none",
        "semantic_owner": "none",
        "runtime_truth": False,
        **NON_CLAIM_FLAGS,
        "pipeline_stage_boundary": {
            "implemented_stage": "raw_source_inventory",
            "next_stages_not_implemented": [
                "candidate",
                "validator_passed",
                "auto_eligible_packet_candidate",
                "packet_ready",
            ],
        },
        "sources": [definition.as_dict() for definition in RAW_SOURCE_DEFINITIONS],
    }


def build_food_raw_source_inventory(scan_roots: Iterable[Path | str]) -> dict[str, Any]:
    roots = [Path(root) for root in scan_roots]
    entries = [
        _inventory_entry(definition, roots)
        for definition in RAW_SOURCE_DEFINITIONS
    ]
    present_count = sum(1 for entry in entries if entry["local_path_present"] is True)
    return {
        "artifact_type": "accurate_intake_food_raw_source_inventory",
        "artifact_schema_version": "1.0",
        "generated_at_utc": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "claim_scope": "raw_source_inventory_only",
        "truth_owner": "none",
        "semantic_owner": "none",
        "runtime_truth": False,
        **NON_CLAIM_FLAGS,
        "pipeline_stage_boundary": {
            "implemented_stage": "raw_source_inventory",
            "next_stages_not_implemented": [
                "candidate",
                "validator_passed",
                "auto_eligible_packet_candidate",
                "packet_ready",
            ],
        },
        "scan_summary": {
            "scan_root_count": len(roots),
            "present_count": present_count,
            "absent_count": len(entries) - present_count,
        },
        "inventory_entries": entries,
    }


def _inventory_entry(definition: RawSourceDefinition, scan_roots: list[Path]) -> dict[str, Any]:
    match = _find_source_file(definition.filename, scan_roots)
    base = definition.as_dict()
    base.update(
        {
            "local_path_present": False,
            "extension": Path(definition.filename).suffix.lower(),
            "file_size": None,
            "path_hash": None,
            "relative_to_scan_root": None,
            "row_count": None,
        }
    )
    if match is None:
        return base

    path, root = match
    base.update(
        {
            "local_path_present": True,
            "file_size": path.stat().st_size,
            "path_hash": _path_hash(path),
            "relative_to_scan_root": _relative_to_root(path, root),
        }
    )
    suffix = path.suffix.lower()
    if suffix == ".json":
        base.update(_inspect_json(path))
    elif suffix == ".csv":
        base.update(_inspect_csv(path))
    elif suffix == ".xlsx":
        base.update(_inspect_xlsx(path))
    return base


def _find_source_file(filename: str, scan_roots: list[Path]) -> tuple[Path, Path] | None:
    for root in scan_roots:
        if not root.exists():
            continue
        if root.is_file() and root.name == filename:
            return root, root.parent
        if root.is_dir():
            direct = root / filename
            if direct.exists():
                return direct, root
            for candidate in root.rglob(filename):
                if candidate.is_file():
                    return candidate, root
    return None


def _inspect_json(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except (json.JSONDecodeError, OSError, UnicodeError) as exc:
        return {"parse_error": type(exc).__name__}

    records = _json_records(payload)
    schema_keys = sorted(
        {
            str(key)
            for record in records
            if isinstance(record, dict)
            for key in record.keys()
        }
    )
    return {
        "row_count": len(records),
        "schema_keys": schema_keys,
        "schema_fingerprint": _schema_fingerprint(schema_keys),
    }


def _inspect_csv(path: Path) -> dict[str, Any]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
    except (csv.Error, OSError, UnicodeError) as exc:
        return {"parse_error": type(exc).__name__}

    schema_keys = sorted(str(name) for name in (reader.fieldnames or []) if name)
    return {
        "row_count": len(rows),
        "schema_keys": schema_keys,
        "schema_fingerprint": _schema_fingerprint(schema_keys),
    }


def _json_records(payload: Any) -> list[Any]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        records = payload.get("records")
        if isinstance(records, list):
            return records
    return []


def _inspect_xlsx(path: Path) -> dict[str, Any]:
    try:
        _preflight_xlsx_worksheet_xml(path)
        workbook = _load_workbook_read_only(path)
    except XLSX_PARSE_EXCEPTIONS as exc:
        return {"parse_error": type(exc).__name__}
    try:
        sheets = []
        row_count = 0
        for sheet in workbook.worksheets:
            header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header_row = [str(value) for value in header[:20] if value is not None]
            data_rows = max((sheet.max_row or 0) - 1, 0)
            row_count += data_rows
            sheets.append(
                {
                    "title": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "header_row": header_row,
                }
            )
        return {
            "row_count": row_count,
            "sheet_count": len(sheets),
            "sheets": sheets,
        }
    except XLSX_PARSE_EXCEPTIONS as exc:
        return {"parse_error": type(exc).__name__}
    finally:
        workbook.close()


def _load_workbook_read_only(path: Path):
    reader = ExcelReader(path, read_only=True, data_only=True)
    try:
        reader.read()
    except XLSX_PARSE_EXCEPTIONS:
        if reader.wb is not None:
            reader.wb.close()
        reader.archive.close()
        raise
    return reader.wb


def _preflight_xlsx_worksheet_xml(path: Path) -> None:
    with ZipFile(path, "r") as archive:
        worksheet_names = [
            name
            for name in archive.namelist()
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        ]
        for name in worksheet_names:
            with archive.open(name) as source:
                for _event, element in iterparse(source):
                    element.clear()


def _schema_fingerprint(schema_keys: list[str]) -> str:
    payload = "\n".join(schema_keys).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _path_hash(path: Path) -> str:
    return hashlib.sha256(str(path.resolve()).encode("utf-8")).hexdigest()


def _relative_to_root(path: Path, root: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return path.name


__all__ = [
    "RAW_SOURCE_DEFINITIONS",
    "build_food_raw_source_inventory",
    "build_food_raw_source_registry",
]
