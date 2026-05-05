from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import ParseError, iterparse
from zipfile import BadZipFile, ZipFile

from openpyxl.reader.excel import ExcelReader
from openpyxl.utils.exceptions import InvalidFileException

XLSX_PARSE_EXCEPTIONS = (
    BadZipFile,
    InvalidFileException,
    ParseError,
    OSError,
    ValueError,
    KeyError,
)


def inspect_json(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except (json.JSONDecodeError, OSError, UnicodeError) as exc:
        return {"parse_error": type(exc).__name__}

    records = json_records(payload)
    schema_keys = sorted(
        {
            str(key)
            for record in records
            if isinstance(record, dict)
            for key in record.keys()
        }
    )
    return {
        "row_count": len(records),
        "schema_keys": schema_keys,
        "schema_fingerprint": schema_fingerprint(schema_keys),
    }


def inspect_csv(path: Path) -> dict[str, Any]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
    except (csv.Error, OSError, UnicodeError) as exc:
        return {"parse_error": type(exc).__name__}

    schema_keys = sorted(str(name) for name in (reader.fieldnames or []) if name)
    return {
        "row_count": len(rows),
        "schema_keys": schema_keys,
        "schema_fingerprint": schema_fingerprint(schema_keys),
    }


def json_records(payload: Any) -> list[Any]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        records = payload.get("records")
        if isinstance(records, list):
            return records
    return []


def inspect_xlsx(path: Path) -> dict[str, Any]:
    try:
        preflight_xlsx_worksheet_xml(path)
        workbook = load_workbook_read_only(path)
    except XLSX_PARSE_EXCEPTIONS as exc:
        return {"parse_error": type(exc).__name__}
    try:
        sheets = []
        row_count = 0
        for sheet in workbook.worksheets:
            header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header_row = [str(value) for value in header[:20] if value is not None]
            data_rows = max((sheet.max_row or 0) - 1, 0)
            row_count += data_rows
            sheets.append(
                {
                    "title": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "header_row": header_row,
                }
            )
        return {
            "row_count": row_count,
            "sheet_count": len(sheets),
            "sheets": sheets,
        }
    except XLSX_PARSE_EXCEPTIONS as exc:
        return {"parse_error": type(exc).__name__}
    finally:
        workbook.close()


def load_workbook_read_only(path: Path):
    reader = ExcelReader(path, read_only=True, data_only=True)
    try:
        reader.read()
    except XLSX_PARSE_EXCEPTIONS:
        if reader.wb is not None:
            reader.wb.close()
        reader.archive.close()
        raise
    return reader.wb


def preflight_xlsx_worksheet_xml(path: Path) -> None:
    with ZipFile(path, "r") as archive:
        worksheet_names = [
            name
            for name in archive.namelist()
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        ]
        for name in worksheet_names:
            with archive.open(name) as source:
                for _event, element in iterparse(source):
                    element.clear()


def schema_fingerprint(schema_keys: list[str]) -> str:
    payload = "\n".join(schema_keys).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


__all__ = [
    "XLSX_PARSE_EXCEPTIONS",
    "inspect_csv",
    "inspect_json",
    "inspect_xlsx",
    "json_records",
    "load_workbook_read_only",
    "preflight_xlsx_worksheet_xml",
    "schema_fingerprint",
]
