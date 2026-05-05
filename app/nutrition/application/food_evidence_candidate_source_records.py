from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from app.nutrition.application.food_evidence_candidate_source_file_records import (
    csv_records,
    find_source_file,
    json_records,
)
from app.nutrition.application.food_evidence_candidate_source_file_records import (
    schema_keys as collect_schema_keys,
)
from app.nutrition.application.food_evidence_candidate_xlsx_records import (
    detect_tfda_header,
    xlsx_row_record,
)
from app.nutrition.application.food_raw_source_inventory import RawSourceDefinition

CandidateResult = tuple[dict[str, Any], list[str]]
RecordNormalizer = Callable[[RawSourceDefinition, dict[str, Any], int], CandidateResult]
RejectionBuilder = Callable[
    [RawSourceDefinition, int, str | None, list[str]], dict[str, Any]
]
RecordIdResolver = Callable[[dict[str, Any]], str | None]
XlsxRecordNormalizer = Callable[
    [RawSourceDefinition, dict[str, Any], int], CandidateResult
]


def normalize_source(
    definition: RawSourceDefinition,
    scan_roots: list[Path],
    *,
    normalize_json_record: RecordNormalizer,
    normalize_csv_record: RecordNormalizer,
    normalize_xlsx_record: XlsxRecordNormalizer,
    build_rejection: RejectionBuilder,
    resolve_record_id: RecordIdResolver,
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    report = _empty_source_report(definition)
    match = find_source_file(definition.filename, scan_roots)
    if match is None:
        return report, [], []

    path = match
    report["local_path_present"] = True
    try:
        if path.suffix.lower() == ".xlsx":
            candidates, rejections, schema_keys, parsed_count = _normalize_xlsx(
                definition, path, normalize_xlsx_record, build_rejection
            )
        elif path.suffix.lower() == ".csv":
            records = csv_records(path)
            schema_keys = collect_schema_keys(records)
            candidates, rejections = _normalize_csv_records(
                definition,
                records,
                normalize_csv_record,
                build_rejection,
                resolve_record_id,
            )
            parsed_count = len(records)
        else:
            payload = json.loads(path.read_text(encoding="utf-8-sig"))
            records = json_records(payload)
            schema_keys = collect_schema_keys(records)
            candidates, rejections = _normalize_json_records(
                definition,
                records,
                normalize_json_record,
                build_rejection,
                resolve_record_id,
            )
            parsed_count = len(records)
    except (csv.Error, json.JSONDecodeError, OSError, UnicodeError) as exc:
        report["parse_error"] = type(exc).__name__
        return report, [], []

    report.update(
        {
            "parsed_count": parsed_count,
            "candidate_count": len(candidates),
            "rejected_count": len(rejections),
            "schema_keys": schema_keys,
        }
    )
    return report, candidates, rejections


def _empty_source_report(definition: RawSourceDefinition) -> dict[str, Any]:
    return {
        "source_id": definition.source_id,
        "filename": definition.filename,
        "source_class": definition.source_class,
        "source_role": definition.source_role,
        "local_path_present": False,
        "parsed_count": 0,
        "candidate_count": 0,
        "rejected_count": 0,
        "schema_keys": [],
        "parse_error": None,
    }


def _normalize_json_records(
    definition: RawSourceDefinition,
    records: list[Any],
    normalize_record: RecordNormalizer,
    build_rejection: RejectionBuilder,
    resolve_record_id: RecordIdResolver,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    candidates: list[dict[str, Any]] = []
    rejections: list[dict[str, Any]] = []
    for index, record in enumerate(records, start=1):
        if not isinstance(record, dict):
            rejections.append(
                build_rejection(definition, index, None, ["invalid_record_shape"])
            )
            continue
        candidate, reasons = normalize_record(definition, record, index)
        if reasons:
            rejections.append(
                build_rejection(definition, index, resolve_record_id(record), reasons)
            )
            continue
        candidates.append(candidate)
    return candidates, rejections


def _normalize_csv_records(
    definition: RawSourceDefinition,
    records: list[Any],
    normalize_record: RecordNormalizer,
    build_rejection: RejectionBuilder,
    resolve_record_id: RecordIdResolver,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    candidates: list[dict[str, Any]] = []
    rejections: list[dict[str, Any]] = []
    for index, record in enumerate(records, start=2):
        if not isinstance(record, dict):
            rejections.append(
                build_rejection(definition, index, None, ["invalid_record_shape"])
            )
            continue
        candidate, reasons = normalize_record(definition, record, index)
        if reasons:
            rejections.append(
                build_rejection(definition, index, resolve_record_id(record), reasons)
            )
            continue
        candidates.append(candidate)
    return candidates, rejections


def _normalize_xlsx(
    definition: RawSourceDefinition,
    path: Path,
    normalize_record: XlsxRecordNormalizer,
    build_rejection: RejectionBuilder,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return [], [], [], 0

    header_index, columns = detect_tfda_header(rows)
    schema_keys = list(columns.keys())
    candidates: list[dict[str, Any]] = []
    rejections: list[dict[str, Any]] = []
    parsed_count = 0

    for zero_index, row in enumerate(rows[header_index + 1 :], start=header_index + 1):
        parsed_count += 1
        record = xlsx_row_record(row, columns)
        row_index = zero_index + 1
        candidate, reasons = normalize_record(definition, record, row_index)
        if reasons:
            rejections.append(build_rejection(definition, row_index, None, reasons))
            continue
        candidates.append(candidate)
    return candidates, rejections, schema_keys, parsed_count


__all__ = [
    "normalize_source",
]
