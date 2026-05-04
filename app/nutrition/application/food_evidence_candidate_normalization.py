from __future__ import annotations

from datetime import UTC, datetime
import hashlib
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from app.nutrition.application.food_raw_source_inventory import (
    NON_CLAIM_FLAGS,
    RAW_SOURCE_DEFINITIONS,
    RawSourceDefinition,
)


NO_TRUTH_FLAGS = {
    **NON_CLAIM_FLAGS,
    "runtime_truth_changed": False,
}

EVIDENCE_ROLE_BY_SOURCE_ID = {
    "newtaipei_brand_candidates": "exact_card_candidate",
    "openfoodfacts_taiwan_small": "packaged_candidate",
    "usda_food_list_sample": "fallback_anchor_candidate",
    "base_nutrition_db": "alias_coverage_prior",
}
TFDA_LISTED_COMPONENT_PREFIXES = (
    "豆干",
    "豆乾",
    "海帶",
    "海帶芽",
    "貢丸",
    "高麗菜",
    "四季豆",
    "青江菜",
    "小白菜",
    "空心菜",
    "A菜",
    "大陸妹",
    "萵苣",
    "花椰菜",
    "豆皮",
    "豆腐皮",
    "百頁豆腐",
    "米血",
    "豬血糕",
    "黑輪",
    "甜不辣",
    "香菇",
    "金針菇",
    "杏鮑菇",
    "鴨血",
    "凍豆腐",
    "白蘿蔔",
    "蘿蔔",
    "玉米筍",
    "木耳",
    "冬粉",
)
TFDA_LISTED_COMPONENT_DISQUALIFIERS = (
    "奶茶",
    "飲料",
    "水餃",
    "壽司",
    "蒸蛋",
    "魚",
    "雞",
    "罐頭",
    "料理包",
    "蘿蔔糕",
    "麵",
    "酥",
)


def build_food_evidence_candidate_artifact(
    scan_roots: Iterable[Path | str],
) -> dict[str, Any]:
    roots = [Path(root) for root in scan_roots]
    candidates: list[dict[str, Any]] = []
    rejections: list[dict[str, Any]] = []
    reports: list[dict[str, Any]] = []

    for definition in RAW_SOURCE_DEFINITIONS:
        report, source_candidates, source_rejections = _normalize_source(definition, roots)
        reports.append(report)
        candidates.extend(source_candidates)
        rejections.extend(source_rejections)

    parse_error_count = sum(1 for report in reports if report.get("parse_error"))
    present_reports = [report for report in reports if report["local_path_present"]]

    return {
        "artifact_type": "accurate_intake_food_evidence_candidates",
        "artifact_schema_version": "1.0",
        "generated_at_utc": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "claim_scope": "food_evidence_candidate_normalization_only",
        "truth_owner": "none",
        "semantic_owner": "none",
        "runtime_truth": False,
        **NO_TRUTH_FLAGS,
        "pipeline_stage_boundary": {
            "implemented_stage": "candidate",
            "next_stages_not_implemented": [
                "validator_passed",
                "auto_eligible_packet_candidate",
                "packet_ready",
            ],
        },
        "candidate_summary": {
            "source_count": len(RAW_SOURCE_DEFINITIONS),
            "present_source_count": len(present_reports),
            "parsed_count": sum(report["parsed_count"] for report in reports),
            "candidate_count": len(candidates),
            "rejected_count": len(rejections),
            "parse_error_count": parse_error_count,
        },
        "source_reports": reports,
        "candidates": candidates,
        "rejections": rejections,
    }


def _normalize_source(
    definition: RawSourceDefinition,
    scan_roots: list[Path],
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    report = _empty_source_report(definition)
    match = _find_source_file(definition.filename, scan_roots)
    if match is None:
        return report, [], []

    path = match
    report["local_path_present"] = True
    try:
        if path.suffix.lower() == ".xlsx":
            candidates, rejections, schema_keys, parsed_count = _normalize_xlsx(definition, path)
        else:
            payload = json.loads(path.read_text(encoding="utf-8-sig"))
            records = _json_records(payload)
            schema_keys = _schema_keys(records)
            candidates, rejections = _normalize_json_records(definition, records)
            parsed_count = len(records)
    except (json.JSONDecodeError, OSError, UnicodeError) as exc:
        report["parse_error"] = type(exc).__name__
        return report, [], []

    report.update(
        {
            "parsed_count": parsed_count,
            "candidate_count": len(candidates),
            "rejected_count": len(rejections),
            "schema_keys": schema_keys,
        }
    )
    return report, candidates, rejections


def _empty_source_report(definition: RawSourceDefinition) -> dict[str, Any]:
    return {
        "source_id": definition.source_id,
        "filename": definition.filename,
        "source_class": definition.source_class,
        "source_role": definition.source_role,
        "local_path_present": False,
        "parsed_count": 0,
        "candidate_count": 0,
        "rejected_count": 0,
        "schema_keys": [],
        "parse_error": None,
    }


def _normalize_json_records(
    definition: RawSourceDefinition,
    records: list[Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    candidates: list[dict[str, Any]] = []
    rejections: list[dict[str, Any]] = []
    for index, record in enumerate(records, start=1):
        if not isinstance(record, dict):
            rejections.append(_rejection(definition, index, None, ["invalid_record_shape"]))
            continue
        candidate, reasons = _candidate_from_json_record(definition, record, index)
        if reasons:
            rejections.append(_rejection(definition, index, _record_id(record), reasons))
            continue
        candidates.append(candidate)
    return candidates, rejections


def _candidate_from_json_record(
    definition: RawSourceDefinition,
    record: dict[str, Any],
    row_index: int,
) -> tuple[dict[str, Any], list[str]]:
    source_id = definition.source_id
    label = _label_for_json_record(source_id, record)
    kcal = _kcal_for_json_record(source_id, record)
    reasons = _basic_rejection_reasons(label=label, kcal=kcal)
    if reasons:
        return {}, reasons

    aliases = _aliases_for_json_record(source_id, record, label)
    serving_basis = _serving_basis_for_json_record(source_id, record)
    return (
        _candidate(
            definition=definition,
            label=label,
            row_index=row_index,
            record_id=_record_id(record),
            kcal=kcal,
            aliases=aliases,
            category=_first_text(record, ("category", "dataType")),
            brand=_brand_for_json_record(source_id, record),
            serving_basis=serving_basis,
            source_url=_first_text(record, ("source_url", "url")),
            raw_record=record,
        ),
        [],
    )


def _normalize_xlsx(
    definition: RawSourceDefinition,
    path: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return [], [], [], 0

    header_index, columns = _detect_tfda_header(rows)
    schema_keys = list(columns.keys())
    candidates: list[dict[str, Any]] = []
    rejections: list[dict[str, Any]] = []
    parsed_count = 0

    for zero_index, row in enumerate(rows[header_index + 1 :], start=header_index + 1):
        parsed_count += 1
        record = _xlsx_row_record(row, columns)
        label = _text(record.get("label"))
        kcal = _number(record.get("corrected_kcal")) or _number(record.get("kcal"))
        reasons = _basic_rejection_reasons(label=label, kcal=kcal)
        row_index = zero_index + 1
        if reasons:
            rejections.append(_rejection(definition, row_index, None, reasons))
            continue
        candidates.append(
            _candidate(
                definition=definition,
                label=label,
                row_index=row_index,
                record_id=None,
                kcal=kcal,
                aliases=_split_aliases(record.get("aliases")),
                category=_text(record.get("category")),
                brand=None,
                serving_basis={
                    "unit_type": "g",
                    "amount": 100,
                    "label": "per_100g_edible_portion",
                },
                source_url=None,
                raw_record=record,
            )
        )
    return candidates, rejections, schema_keys, parsed_count


def _candidate(
    *,
    definition: RawSourceDefinition,
    label: str,
    row_index: int,
    record_id: str | None,
    kcal: float,
    aliases: list[str],
    category: str | None,
    brand: str | None,
    serving_basis: dict[str, Any],
    source_url: str | None,
    raw_record: dict[str, Any],
) -> dict[str, Any]:
    candidate_hash = _stable_hash(
        {
            "source_id": definition.source_id,
            "label": label,
            "row_index": row_index,
            "record_id": record_id,
            "kcal": kcal,
        }
    )[:12]
    return {
        "candidate_id": f"cand_{definition.source_id}_{candidate_hash}",
        "source_id": definition.source_id,
        "source_class": definition.source_class,
        "source_role": definition.source_role,
        "evidence_role": _evidence_role(definition, label=label, aliases=aliases),
        "promotion_status": "candidate",
        "runtime_truth_allowed": False,
        "canonical_label": label,
        "aliases": aliases,
        "brand": brand,
        "category": category,
        "serving_basis": serving_basis,
        "kcal_point": kcal,
        "kcal_range": None,
        "source_provenance": {
            "source_id": definition.source_id,
            "source_file": definition.filename,
            "row_index": row_index,
            "record_id": record_id,
            "source_url": source_url,
            "raw_row_hash": _stable_hash(raw_record),
        },
        "quality_flags": [],
    }


def _rejection(
    definition: RawSourceDefinition,
    row_index: int,
    record_id: str | None,
    reasons: list[str],
) -> dict[str, Any]:
    return {
        "source_id": definition.source_id,
        "source_class": definition.source_class,
        "source_file": definition.filename,
        "row_index": row_index,
        "record_id": record_id,
        "reasons": reasons,
    }


def _find_source_file(filename: str, scan_roots: list[Path]) -> Path | None:
    for root in scan_roots:
        if not root.exists():
            continue
        if root.is_file() and root.name == filename:
            return root
        if root.is_dir():
            direct = root / filename
            if direct.exists():
                return direct
            for candidate in root.rglob(filename):
                if candidate.is_file():
                    return candidate
    return None


def _json_records(payload: Any) -> list[Any]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        records = payload.get("records")
        if isinstance(records, list):
            return records
        products = payload.get("products")
        if isinstance(products, list):
            return products
    return []


def _schema_keys(records: list[Any]) -> list[str]:
    return sorted(
        {
            str(key)
            for record in records
            if isinstance(record, dict)
            for key in record.keys()
        }
    )


def _label_for_json_record(source_id: str, record: dict[str, Any]) -> str:
    if source_id == "base_nutrition_db":
        return _first_text(record, ("title", "name", "label"))
    if source_id == "openfoodfacts_taiwan_small":
        return _first_text(record, ("product_name", "generic_name"))
    if source_id == "usda_food_list_sample":
        return _first_text(record, ("description", "name"))
    if source_id == "tfda_base_candidates":
        return _first_text(record, ("variant", "title", "name"))
    return _first_text(record, ("title", "variant", "name", "label"))


def _aliases_for_json_record(
    source_id: str,
    record: dict[str, Any],
    label: str,
) -> list[str]:
    raw_aliases: list[Any] = []
    if source_id == "base_nutrition_db":
        aliases = record.get("aliases")
        if isinstance(aliases, list):
            raw_aliases.extend(aliases)
        elif aliases:
            raw_aliases.append(aliases)
    if source_id in {"tfda_base_candidates", "tfda_base_review_candidates"}:
        raw_aliases.append(record.get("brand"))
        variant = _text(record.get("variant"))
        title = _text(record.get("title"))
        if variant and variant != label:
            raw_aliases.append(variant)
        if title and title != label:
            raw_aliases.append(title)
    if source_id == "openfoodfacts_taiwan_small":
        raw_aliases.append(record.get("generic_name"))
    return _dedupe(
        alias
        for raw in raw_aliases
        for alias in _split_aliases(raw)
        if alias and alias != label
    )


def _brand_for_json_record(source_id: str, record: dict[str, Any]) -> str | None:
    if source_id == "openfoodfacts_taiwan_small":
        return _first_text(record, ("brands", "brand"))
    if source_id == "newtaipei_brand_candidates":
        return _first_text(record, ("brand", "brands"))
    return None


def _serving_basis_for_json_record(source_id: str, record: dict[str, Any]) -> dict[str, Any]:
    basis = record.get("serving_basis")
    if isinstance(basis, dict):
        return basis
    if source_id in {"openfoodfacts_taiwan_small", "usda_food_list_sample"}:
        return {"unit_type": "g", "amount": 100, "label": "per_100g"}
    return {"unit_type": "g", "amount": 100, "label": "per_100g"}


def _kcal_for_json_record(source_id: str, record: dict[str, Any]) -> float | None:
    if source_id == "base_nutrition_db":
        nutrition = record.get("nutrition")
        if isinstance(nutrition, dict):
            return _number(nutrition.get("kcal"))
    if source_id == "openfoodfacts_taiwan_small":
        nutriments = record.get("nutriments")
        if isinstance(nutriments, dict):
            return _number(
                nutriments.get("energy-kcal_100g")
                or nutriments.get("energy-kcal")
                or nutriments.get("energy_kcal_100g")
            )
    if source_id == "usda_food_list_sample":
        nutrients = record.get("foodNutrients")
        if isinstance(nutrients, list):
            for nutrient in nutrients:
                if not isinstance(nutrient, dict):
                    continue
                nutrient_name = str(nutrient.get("nutrientName") or "").lower()
                unit_name = str(nutrient.get("unitName") or "").upper()
                if "energy" in nutrient_name and unit_name == "KCAL":
                    return _number(nutrient.get("value"))
    return _number(record.get("kcal") or record.get("calories") or record.get("energy_kcal"))


def _record_id(record: dict[str, Any]) -> str | None:
    value = (
        record.get("id")
        or record.get("code")
        or record.get("fdcId")
        or record.get("source_id")
    )
    if value is None:
        return None
    return str(value)


def _basic_rejection_reasons(label: str, kcal: float | None) -> list[str]:
    reasons: list[str] = []
    if not label:
        reasons.append("missing_label")
    if kcal is None:
        reasons.append("missing_kcal")
    return reasons


def _evidence_role(
    definition: RawSourceDefinition,
    *,
    label: str,
    aliases: list[str],
) -> str:
    if definition.source_id in EVIDENCE_ROLE_BY_SOURCE_ID:
        return EVIDENCE_ROLE_BY_SOURCE_ID[definition.source_id]
    if definition.source_class == "taiwan_tfda_open_data":
        if _is_tfda_listed_component_candidate(label=label, aliases=aliases):
            return "listed_component_anchor_candidate"
        return "generic_anchor_candidate"
    return f"{definition.intended_roles[0]}_candidate"


def _is_tfda_listed_component_candidate(*, label: str, aliases: list[str]) -> bool:
    tokens = [label] if label else list(aliases)
    for token in tokens:
        normalized = str(token or "").strip()
        if not normalized:
            continue
        if any(disqualifier in normalized for disqualifier in TFDA_LISTED_COMPONENT_DISQUALIFIERS):
            continue
        if normalized.endswith("乾") and not normalized.startswith(("豆干", "豆乾")):
            continue
        if any(normalized.startswith(prefix) for prefix in TFDA_LISTED_COMPONENT_PREFIXES):
            return True
    return False


def _detect_tfda_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows[:8]):
        values = [_text(value) for value in row]
        if "樣品名稱" in values and ("修正熱量(kcal)" in values or "熱量(kcal)" in values):
            return index, {
                "category": values.index("食品分類") if "食品分類" in values else 0,
                "label": values.index("樣品名稱"),
                "aliases": values.index("俗名") if "俗名" in values else 2,
                "kcal": values.index("熱量(kcal)") if "熱量(kcal)" in values else 3,
                "corrected_kcal": values.index("修正熱量(kcal)")
                if "修正熱量(kcal)" in values
                else 4,
            }
    return 1, {
        "category": 0,
        "label": 1,
        "aliases": 2,
        "kcal": 3,
        "corrected_kcal": 4,
    }


def _xlsx_row_record(row: tuple[Any, ...], columns: dict[str, int]) -> dict[str, Any]:
    return {
        field: row[index] if index < len(row) else None
        for field, index in columns.items()
    }


def _first_text(record: dict[str, Any], keys: tuple[str, ...]) -> str:
    for key in keys:
        value = _text(record.get(key))
        if value:
            return value
    return ""


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _split_aliases(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return _dedupe(alias for item in value for alias in _split_aliases(item))
    text = str(value).strip()
    if not text:
        return []
    for delimiter in ("、", "，", ",", ";", "；", "/"):
        text = text.replace(delimiter, "\n")
    return _dedupe(part.strip() for part in text.splitlines() if part.strip())


def _dedupe(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def _stable_hash(payload: Any) -> str:
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


__all__ = ["build_food_evidence_candidate_artifact"]
