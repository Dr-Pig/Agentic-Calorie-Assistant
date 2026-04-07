from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_KNOWLEDGE = ROOT / "app" / "knowledge"
DEFAULT_CANDIDATE_PATH = ROOT / ".logs" / "tfda_base_candidates_tmp.json"
DEFAULT_WORKBOOK_PATH = ROOT / ".logs" / "FDA_food_nutrition_2024.xlsx"
DEFAULT_OUTPUT_PATH = ROOT / ".logs" / "tfda_base_enriched_candidates.json"


TFDA_SOURCE_URL = "https://consumer.fda.gov.tw/Food/TFND.aspx?nodeID=178"
TFDA_SOURCE_NAME = "Taiwan TFDA Food Nutrition Database Portal"
HEADER_ROW_INDEX = 2


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _safe_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _safe_text(value))
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _normalize_key(value: Any) -> str:
    text = _normalize_text(value).lower()
    text = re.sub(r"[;,/()\-]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _slugify(value: str) -> str:
    text = _normalize_key(value)
    text = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "-", text)
    text = re.sub(r"-{2,}", "-", text)
    return text.strip("-") or "tfda-item"


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _kcal_round_key(value: Any) -> int | None:
    numeric = _to_float(value)
    if numeric is None:
        return None
    return int(round(numeric))


def _alias_tokens(text: str) -> list[str]:
    normalized = _normalize_text(text)
    if not normalized:
        return []
    parts = re.split(r"[,，、/]+", normalized)
    return [part.strip() for part in parts if part and part.strip()]


def _build_aliases(candidate: dict[str, Any], row: dict[str, Any]) -> list[str]:
    aliases: list[str] = []
    candidate_brand = _safe_text(candidate.get("brand"))
    brand_values: list[str] = []
    if candidate_brand and candidate_brand not in {"TFDA營養資料", "TFDA", "government_dataset"}:
        brand_values.append(candidate_brand)
    for value in [
        row.get("樣品名稱"),
        row.get("俗名"),
        candidate.get("variant"),
        *brand_values,
    ]:
        for token in _alias_tokens(_safe_text(value)):
            if token and token not in aliases:
                aliases.append(token)
    return aliases


def _load_tfda_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_values = list(next(worksheet.iter_rows(min_row=HEADER_ROW_INDEX, max_row=HEADER_ROW_INDEX, values_only=True)))
    rows: list[dict[str, Any]] = []
    for raw_row in worksheet.iter_rows(min_row=HEADER_ROW_INDEX + 1, values_only=True):
        if not raw_row or not any(raw_row):
            continue
        row = {str(header_values[idx]).strip(): raw_row[idx] for idx in range(min(len(header_values), len(raw_row))) if header_values[idx]}
        if _safe_text(row.get("整合編號")):
            rows.append(row)
    return rows


def _build_row_index(rows: list[dict[str, Any]]) -> dict[tuple[str, str, str, int | None], dict[str, Any]]:
    index: dict[tuple[str, str, str, int | None], dict[str, Any]] = {}
    for row in rows:
        key = (
            _normalize_key(row.get("食品分類")),
            _normalize_key(row.get("內容物描述")),
            _normalize_key(row.get("樣品名稱")),
            _kcal_round_key(row.get("修正熱量(kcal)")) or _kcal_round_key(row.get("熱量(kcal)")),
        )
        index[key] = row
    return index


def _build_row_groups(rows: list[dict[str, Any]]) -> dict[tuple[str, str, str], list[dict[str, Any]]]:
    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    for row in rows:
        key = (
            _normalize_key(row.get("食品分類")),
            _normalize_key(row.get("內容物描述")),
            _normalize_key(row.get("樣品名稱")),
        )
        groups.setdefault(key, []).append(row)
    return groups


def _match_row(
    candidate: dict[str, Any],
    row_index: dict[tuple[str, str, str, int | None], dict[str, Any]],
    row_groups: dict[tuple[str, str, str], list[dict[str, Any]]],
) -> dict[str, Any] | None:
    kcal_key = _kcal_round_key(candidate.get("kcal"))
    triple_key = (
        _normalize_key(candidate.get("category")),
        _normalize_key(candidate.get("title")),
        _normalize_key(candidate.get("variant")),
    )
    keys = [
        (*triple_key, kcal_key),
        (*triple_key, None),
    ]
    for key in keys:
        if key in row_index:
            return row_index[key]
    candidates = row_groups.get(triple_key, [])
    if not candidates:
        return None
    target_kcal = _to_float(candidate.get("kcal"))
    if target_kcal is None:
        return candidates[0]
    best_row: dict[str, Any] | None = None
    best_diff: float | None = None
    for row in candidates:
        row_kcal = _to_float(row.get("修正熱量(kcal)")) or _to_float(row.get("熱量(kcal)"))
        if row_kcal is None:
            continue
        diff = abs(row_kcal - target_kcal)
        if best_diff is None or diff < best_diff:
            best_diff = diff
            best_row = row
    if best_row is not None and best_diff is not None and best_diff <= 1.5:
        return best_row
    return None


def _map_tfda_candidate(candidate: dict[str, Any], row: dict[str, Any]) -> dict[str, Any]:
    item_name = _normalize_text(row.get("樣品名稱"))
    description = _normalize_text(row.get("內容物描述"))
    aliases = _build_aliases(candidate, row)
    sodium_mg = _to_float(row.get("鈉(mg)"))
    kcal = _to_float(row.get("修正熱量(kcal)")) or _to_float(row.get("熱量(kcal)")) or _to_float(candidate.get("kcal")) or 0.0
    notes_parts = [
        f"TFDA integrated id: {_safe_text(row.get('整合編號'))}",
        f"Content description: {description}" if description else "",
        f"Waste rate: {_safe_text(row.get('廢棄率(%)'))}%" if _safe_text(row.get("廢棄率(%)")) else "",
    ]
    return {
        "id": _slugify(f"tfda-{item_name}-{_safe_text(row.get('整合編號'))}"),
        "title": item_name,
        "aliases": aliases,
        "category": _normalize_text(row.get("食品分類")),
        "serving_basis": {
            "unit_type": "g",
            "amount": 100,
            "label": "100 g edible portion",
        },
        "nutrition": {
            "protein_g": _to_float(row.get("粗蛋白(g)")) or 0.0,
            "carb_g": _to_float(row.get("總碳水化合物(g)")) or 0.0,
            "fat_g": _to_float(row.get("粗脂肪(g)")) or 0.0,
            "kcal": kcal,
            "sodium_mg": sodium_mg,
        },
        "portion_equivalents": [
            {
                "label": "100 g",
                "grams": 100,
                "ml": None,
                "pieces": None,
            }
        ],
        "source_type": "government_nutrition",
        "source_name": TFDA_SOURCE_NAME,
        "source_url": TFDA_SOURCE_URL,
        "confidence": "high",
        "last_verified_at": "2026-03-31",
        "notes": " | ".join(part for part in notes_parts if part),
    }


def _merge_records(existing: list[dict[str, Any]], new_records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    merged = deepcopy(existing)
    index = {str(item.get("id")): idx for idx, item in enumerate(merged)}
    inserted = 0
    for record in new_records:
        record_id = str(record.get("id"))
        if record_id in index:
            merged[index[record_id]] = record
            continue
        merged.append(record)
        index[record_id] = len(merged) - 1
        inserted += 1
    return merged, inserted


def build_tfda_base(
    *,
    candidate_path: Path = DEFAULT_CANDIDATE_PATH,
    workbook_path: Path = DEFAULT_WORKBOOK_PATH,
    write_runtime: bool = False,
    output_path: Path | None = DEFAULT_OUTPUT_PATH,
) -> dict[str, Any]:
    candidates_payload = _load_json(candidate_path)
    rows = _load_tfda_rows(workbook_path)
    row_index = _build_row_index(rows)
    row_groups = _build_row_groups(rows)

    enriched_records: list[dict[str, Any]] = []
    unmatched_candidates: list[dict[str, Any]] = []

    for candidate in candidates_payload.get("records", []):
        row = _match_row(candidate, row_index, row_groups)
        if row is None:
            unmatched_candidates.append(
                {
                    "id": candidate.get("id"),
                    "title": candidate.get("title"),
                    "variant": candidate.get("variant"),
                    "category": candidate.get("category"),
                    "kcal": candidate.get("kcal"),
                }
            )
            continue
        enriched_records.append(_map_tfda_candidate(candidate, row))

    report: dict[str, Any] = {
        "candidate_path": str(candidate_path),
        "workbook_path": str(workbook_path),
        "records_seen": len(candidates_payload.get("records", [])),
        "records_enriched": len(enriched_records),
        "records_unmatched": len(unmatched_candidates),
        "unmatched_candidates": unmatched_candidates[:20],
    }

    if output_path is not None:
        payload = {
            "schema_version": "tfda-base-enriched.v1",
            "generated_at": "2026-03-31",
            "records": enriched_records,
        }
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        report["output_path"] = str(output_path)

    if write_runtime:
        base_path = APP_KNOWLEDGE / "base_nutrition_db.json"
        base_payload = _load_json(base_path)
        merged_records, inserted = _merge_records(base_payload.get("records", []), enriched_records)
        base_payload["schema_version"] = "base-nutrition.curated.v1"
        base_payload["derived_from_run"] = "base-nutrition-v2-2-20260330-094105 + tfda-base-enrichment-20260331"
        base_payload["reviewed_at"] = "2026-03-31"
        base_payload["records"] = merged_records
        base_path.write_text(json.dumps(base_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        report["runtime_base_count"] = len(merged_records)
        report["runtime_inserted_or_updated"] = inserted

    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Build TFDA macro-complete base nutrition records from staging candidates and TFDA workbook.")
    parser.add_argument("--candidate-path", default=str(DEFAULT_CANDIDATE_PATH), help="Path to tfda_base_candidates JSON.")
    parser.add_argument("--workbook-path", default=str(DEFAULT_WORKBOOK_PATH), help="Path to TFDA workbook.")
    parser.add_argument("--output-path", default=str(DEFAULT_OUTPUT_PATH), help="Where to write enriched candidate JSON.")
    parser.add_argument("--report-out", help="Optional path to save the report JSON.")
    parser.add_argument("--write-runtime", action="store_true", help="Merge enriched records into app/knowledge/base_nutrition_db.json.")
    args = parser.parse_args()

    report = build_tfda_base(
        candidate_path=Path(args.candidate_path).resolve(),
        workbook_path=Path(args.workbook_path).resolve(),
        write_runtime=args.write_runtime,
        output_path=Path(args.output_path).resolve() if args.output_path else None,
    )
    sys.stdout.buffer.write((json.dumps(report, ensure_ascii=False, indent=2) + "\n").encode("utf-8"))
    if args.report_out:
        report_path = Path(args.report_out).resolve()
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
